"""
解析器基类 - 使用 openpyxl 直接处理 Excel
"""
from abc import ABC, abstractmethod
from typing import List, Dict, Optional
from openpyxl import load_workbook
import csv


class BaseParser(ABC):
    """交易流水解析器基类"""

    # 标准化的列名
    STANDARD_COLUMNS = ['date', 'category', 'merchant', 'description', 'amount', 'type', 'source']

    def __init__(self):
        self.source_name = 'unknown'

    @abstractmethod
    def parse(self, file_path: str) -> List[Dict]:
        """
        解析交易流水文件

        Args:
            file_path: 文件路径

        Returns:
            标准化的数据列表，每个元素是一个字典
        """
        pass

    @abstractmethod
    def can_parse(self, file_path: str) -> bool:
        """
        判断是否能解析该文件
        """
        pass

    def _read_excel(self, file_path: str) -> List[List]:
        """读取 Excel 文件，返回二维列表"""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            wb.close()
            return data
        except Exception as e:
            raise ValueError(f"读取 Excel 文件失败: {e}")

    def _read_csv(self, file_path: str) -> List[List]:
        """读取 CSV 文件，自动检测编码"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']

        for encoding in encodings:
            try:
                data = []
                with open(file_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    for row in reader:
                        data.append(row)
                return data
            except UnicodeDecodeError:
                continue
            except Exception as e:
                raise ValueError(f"读取 CSV 文件失败: {e}")

        raise ValueError("无法识别文件编码，请尝试手动转换编码")

    def _clean_amount(self, value) -> float:
        """清理金额字符串，转换为浮点数"""
        if value is None:
            return 0.0

        if isinstance(value, (int, float)):
            return abs(float(value))

        # 移除货币符号和逗号
        value = str(value).replace('¥', '').replace('￥', '').replace(',', '').strip()

        try:
            return abs(float(value))
        except ValueError:
            return 0.0

    def _standardize_type(self, type_str: str) -> str:
        """标准化交易类型"""
        if not type_str:
            return '支出'

        type_str = str(type_str).strip()

        # 收入关键词
        income_keywords = ['收入', '收款', '退款', '转入', '贷']
        # 支出关键词
        expense_keywords = ['支出', '付款', '消费', '转出', '借']

        for keyword in income_keywords:
            if keyword in type_str:
                return '收入'

        for keyword in expense_keywords:
            if keyword in type_str:
                return '支出'

        return '支出'

    def _clean_data(self, data: List[Dict]) -> List[Dict]:
        """清理数据，移除无效记录"""
        cleaned = []
        for row in data:
            # 移除日期为空的记录
            if row.get('date'):
                cleaned.append(row)
        return cleaned
