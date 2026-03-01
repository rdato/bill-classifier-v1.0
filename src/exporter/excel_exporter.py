"""
Excel 导出器 - 使用 openpyxl
"""
import os
from typing import Dict, List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel 导出器"""

    def __init__(self):
        self.title_font = Font(bold=True, size=12)
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_font_white = Font(bold=True, size=11, color='FFFFFF')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.alignment = Alignment(horizontal='center', vertical='center')

    def export(self, records: List[Dict], output_path: str,
               include_summary: bool = True,
               include_detail: bool = True) -> str:
        """导出数据到 Excel 文件"""
        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        # 创建工作簿
        wb = Workbook()

        # 默认删除第一个空工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 生成分类汇总
        if include_summary:
            summary_data = self._generate_summary(records)
            ws_summary = wb.create_sheet('分类汇总', 0)
            self._write_summary(ws_summary, summary_data)

        # 生成交易明细
        if include_detail:
            detail_records = self._prepare_detail(records)
            ws_detail = wb.create_sheet('交易明细', 1 if include_summary else 0)
            self._write_detail(ws_detail, detail_records)

        # 保存文件
        wb.save(output_path)
        return output_path

    def _generate_summary(self, records: List[Dict]) -> Dict:
        """生成分类汇总数据"""
        summary = {}

        for record in records:
            category = record.get('category', '其他支出')
            trans_type = record.get('type', '支出')
            amount = float(record.get('amount', 0))

            if category not in summary:
                summary[category] = {'收入': 0, '支出': 0, '收入笔数': 0, '支出笔数': 0}

            summary[category][trans_type] = summary[category].get(trans_type, 0) + amount
            count_key = f'{trans_type}笔数'
            summary[category][count_key] = summary[category].get(count_key, 0) + 1

        return summary

    def _write_summary(self, ws, summary_data: Dict):
        """写入汇总表"""
        # 写入标题
        ws.merge_cells('A1:F1')
        ws['A1'] = '分类汇总'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.alignment

        # 写入表头
        headers = ['类别', '支出金额', '收入金额', '净支出', '支出笔数', '收入笔数']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.alignment

        # 写入数据
        row = 4
        total_expense = 0
        total_income = 0
        total_expense_count = 0
        total_income_count = 0

        for category, data in sorted(summary_data.items()):
            expense = data.get('支出', 0)
            income = data.get('收入', 0)
            net = expense - income
            expense_count = data.get('支出笔数', 0)
            income_count = data.get('收入笔数', 0)

            ws.cell(row=row, column=1, value=category).border = self.border
            ws.cell(row=row, column=2, value=expense).border = self.border
            ws.cell(row=row, column=3, value=income).border = self.border
            ws.cell(row=row, column=4, value=net).border = self.border
            ws.cell(row=row, column=5, value=expense_count).border = self.border
            ws.cell(row=row, column=6, value=income_count).border = self.border

            # 格式化数字
            for col in [2, 3, 4]:
                ws.cell(row=row, column=col).number_format = '#,##0.00'
                ws.cell(row=row, column=col).alignment = self.alignment

            total_expense += expense
            total_income += income
            total_expense_count += expense_count
            total_income_count += income_count
            row += 1

        # 写入合计行
        ws.cell(row=row, column=1, value='合计').border = self.border
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_expense).border = self.border
        ws.cell(row=row, column=3, value=total_income).border = self.border
        ws.cell(row=row, column=4, value=total_expense - total_income).border = self.border
        ws.cell(row=row, column=5, value=total_expense_count).border = self.border
        ws.cell(row=row, column=6, value=total_income_count).border = self.border

        for col in [2, 3, 4]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
            ws.cell(row=row, column=col).font = Font(bold=True)

        # 调整列宽
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _prepare_detail(self, records: List[Dict]) -> List[Dict]:
        """准备交易明细"""
        # 按日期排序
        sorted_records = sorted(records, key=lambda x: str(x.get('date', '')), reverse=True)
        return sorted_records

    def _write_detail(self, ws, records: List[Dict]):
        """写入明细表"""
        # 写入标题
        ws.merge_cells('A1:G1')
        ws['A1'] = '交易明细'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.alignment

        # 写入表头
        headers = ['日期', '类别', '交易对方', '描述', '金额', '类型', '来源']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.alignment

        # 写入数据
        for row, record in enumerate(records, 4):
            ws.cell(row=row, column=1, value=record.get('date', '')).border = self.border
            ws.cell(row=row, column=2, value=record.get('category', '')).border = self.border
            ws.cell(row=row, column=3, value=record.get('merchant', '')).border = self.border
            ws.cell(row=row, column=4, value=record.get('description', '')).border = self.border

            amount_cell = ws.cell(row=row, column=5, value=record.get('amount', 0))
            amount_cell.border = self.border
            amount_cell.number_format = '#,##0.00'

            ws.cell(row=row, column=6, value=record.get('type', '')).border = self.border
            ws.cell(row=row, column=7, value=record.get('source', '')).border = self.border

        # 调整列宽
        widths = [12, 12, 20, 30, 12, 8, 10]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
